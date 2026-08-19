"""Importa contratti e svincoli dal database del vecchio gestionale.

Uso:
    .venv\\Scripts\\python.exe strumenti\\importa_da_db.py --prova
    .venv\\Scripts\\python.exe strumenti\\importa_da_db.py

Legge `contracts`, `players` e `contract_releases` da fantacalcio.db e
riscrive i fogli Contratti e Svincoli di dati/Gestione.xlsx.

ATTENZIONE: sostituisce il contenuto dei due fogli, non lo aggiunge. Serve a
travasare una volta i dati dal vecchio progetto; dopo di che la fonte di
verità è l'Excel, e il database non va più toccato.

Vengono importati solo i giocatori che hanno almeno un contratto: nel DB del
gestionale erano rimasti anche diversi record di prova senza contratto.
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Manca openpyxl:  .venv\\Scripts\\python.exe -m pip install openpyxl")

RADICE = Path(__file__).resolve().parent.parent
XLSX = RADICE / "dati" / "Gestione.xlsx"
DB_PREDEFINITO = Path(
    r"d:\Codice\Fanta_Riconferme\fantacalcio\database\fantacalcio.db"
)


def anno_di(data) -> int | None:
    """Da '1996-01-01' ricava 1996. Nel DB conta solo l'anno."""
    if not data:
        return None
    testo = str(data).strip()
    for pezzo in (testo[:4],):
        try:
            anno = int(pezzo)
            if 1900 < anno < 2100:
                return anno
        except ValueError:
            pass
    return None


def scrivi_foglio(wb, nome, intestazioni, righe):
    ws = wb[nome]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for i, testo in enumerate(intestazioni, start=1):
        ws.cell(row=1, column=i, value=testo)
    for r in righe:
        ws.append(r)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa dal DB del gestionale")
    parser.add_argument("--db", type=Path, default=DB_PREDEFINITO)
    parser.add_argument("--prova", action="store_true",
                        help="mostra cosa importerebbe senza salvare")
    args = parser.parse_args()

    if not args.db.exists():
        print(f"ERRORE: non trovo il database {args.db}")
        return 1
    if not XLSX.exists():
        print(f"ERRORE: non trovo {XLSX}")
        return 1

    conn = sqlite3.connect(str(args.db))
    conn.row_factory = sqlite3.Row

    # --- Controllo che gli ID squadra coincidano -------------------------
    wb = openpyxl.load_workbook(XLSX)
    squadre_xl = {}
    for riga in wb["AnagraficaSquadre"].iter_rows(min_row=2, values_only=True):
        if riga[0] is not None:
            squadre_xl[int(riga[0])] = str(riga[2] or "").strip()

    disallineate = []
    for r in conn.execute("SELECT id, name FROM teams"):
        atteso = squadre_xl.get(r["id"])
        if atteso is None or atteso != r["name"]:
            disallineate.append((r["id"], r["name"], atteso))

    if disallineate:
        print("ERRORE: gli ID delle squadre non coincidono fra DB ed Excel.")
        print("Importare adesso assegnerebbe i contratti alle squadre sbagliate.\n")
        for idx, nel_db, nel_xl in disallineate:
            print(f"   ID {idx}: DB '{nel_db}'  ≠  Excel '{nel_xl}'")
        return 1
    print(f"Squadre allineate ({len(squadre_xl)} su entrambi i lati).")

    # --- Contratti --------------------------------------------------------
    contratti = []
    senza_nascita = []
    for r in conn.execute(
        "SELECT p.name AS nome, p.birth_date AS nascita, c.team_id, "
        "       c.start_year, c.buy_price, c.is_active "
        "FROM contracts c JOIN players p ON p.id = c.player_id "
        "ORDER BY c.team_id, p.name"
    ):
        anno_nascita = anno_di(r["nascita"])
        if anno_nascita is None:
            senza_nascita.append(r["nome"])
        contratti.append([
            r["nome"], anno_nascita, r["team_id"],
            r["start_year"], r["buy_price"], 1 if r["is_active"] else 0,
        ])

    # --- Svincoli ---------------------------------------------------------
    svincoli = []
    try:
        for r in conn.execute(
            "SELECT p.name AS nome, r.team_id, r.release_year, "
            "       r.penalty_cost, r.notes "
            "FROM contract_releases r JOIN players p ON p.id = r.player_id "
            "ORDER BY r.release_year, p.name"
        ):
            svincoli.append([r["nome"], r["team_id"], r["release_year"],
                             r["penalty_cost"], r["notes"] or ""])
    except sqlite3.OperationalError as exc:
        # Le colonne di contract_releases potrebbero chiamarsi diversamente
        print(f"   (svincoli non letti: {exc})")

    conn.close()

    # --- Riepilogo --------------------------------------------------------
    print(f"\nContratti trovati: {len(contratti)}")
    per_squadra: dict[int, int] = {}
    for c in contratti:
        per_squadra[c[2]] = per_squadra.get(c[2], 0) + 1
    for idx in sorted(per_squadra):
        print(f"   {squadre_xl[idx]:<28} {per_squadra[idx]} contratti")

    print(f"\n{'Giocatore':<22}{'Nato':>6}{'Sq.':>5}{'Dal':>7}"
          f"{'Prezzo':>8}{'Attivo':>8}")
    print("-" * 56)
    for nome, nascita, team, inizio, prezzo, attivo in contratti:
        print(f"{nome[:21]:<22}{nascita or '?':>6}{team:>5}{inizio:>7}"
              f"{prezzo:>8}{attivo:>8}")

    print(f"\nSvincoli trovati: {len(svincoli)}")
    for s in svincoli:
        print(f"   {s[0]} — squadra {s[1]}, anno {s[2]}, penale {s[3]}")

    if senza_nascita:
        print(f"\n! Senza anno di nascita: {', '.join(senza_nascita)}")
        print("  Verranno trattati come categoria A. Correggili nell'Excel.")

    if args.prova:
        print("\n[PROVA] Niente salvato.")
        return 0

    if not contratti:
        print("\nNessun contratto da importare: non tocco l'Excel.")
        return 0

    scrivi_foglio(
        wb, "Contratti",
        ["Giocatore", "AnnoNascita", "IDSquadra", "AnnoInizio",
         "PrezzoAcquisto", "Attivo"],
        contratti,
    )
    if svincoli:
        scrivi_foglio(
            wb, "Svincoli",
            ["Giocatore", "IDSquadra", "AnnoSvincolo", "Penale", "Note"],
            svincoli,
        )

    wb.save(XLSX)
    print(f"\nScritti {len(contratti)} contratti in {XLSX}")
    print("Ora lancia aggiorna.bat per portarli sul sito.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
