"""Compila il foglio CreditiAnnoNuovo applicando le regole del regolamento.

Uso:
    .venv\\Scripts\\python.exe strumenti\\calcola_crediti.py
    .venv\\Scripts\\python.exe strumenti\\calcola_crediti.py --anno 2027
    .venv\\Scripts\\python.exe strumenti\\calcola_crediti.py --prova   (non salva)

Legge Classifiche, Coppa e CreditiAsta, calcola le tre colonne dei bonus e il
budget della stagione nuova, e li scrive nel foglio CreditiAnnoNuovo.

    budget = 1000 + residui + classifica + miglior punteggio + coppa
    poi limitato fra 900 e 1100

I valori sono quelli di modules/economia.py del gestionale (DEFAULT_RULES).
Se un domani cambiate le regole, si aggiornano le costanti qui sotto.

NOTA: il RESET (partire da 950 invece che da 1000) non è previsto da questo
script, perché nell'Excel non c'è una colonna che lo registri. Se una squadra
ha fatto reset, correggi la sua riga a mano dopo aver lanciato lo script.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Manca openpyxl:  .venv\\Scripts\\python.exe -m pip install openpyxl")

RADICE = Path(__file__).resolve().parent.parent
XLSX = RADICE / "dati" / "Gestione.xlsx"

# --- Regole economiche -------------------------------------------------------
#
# Non sono costanti: si leggono dal foglio RegoleCrediti, come nel gestionale
# le si leggeva dalla tabella `credit_rules`. Così un premio si cambia
# nell'Excel e non in questo file.
#
# Questi valori servono solo come rete di sicurezza se il foglio manca o è
# incompleto, e corrispondono a DEFAULT_RULES di modules/economia.py.
PREDEFINITE = {
    "base": {1: 1000},
    "reset": {1: 950},
    "tetto_max": {1: 1100},
    "tetto_min": {1: 900},
    "riparazione": {1: 50},
    "best_score": {1: 10},
    # Il bonus coppa è per POSIZIONE: solo chi vince prende 10.
    "coppa": {1: 10},
    "classifica": {
        1: 25, 2: 15, 3: 10, 4: 5, 5: 0,
        6: 0, 7: -5, 8: -10, 9: -15, 10: -20,
    },
}


def intero(valore, predefinito=None):
    if valore is None or (isinstance(valore, str) and not valore.strip()):
        return predefinito
    try:
        return int(float(valore))
    except (TypeError, ValueError):
        return predefinito


def carica_regole(wb, avvisi):
    """Legge il foglio RegoleCrediti: categoria -> {posizione: valore}.

    Una posizione vuota vale 1, come fa `get_effective_rules()` del gestionale
    con `posizione or 1`: è per questo che il bonus coppa finisce su {1: 10} e
    il secondo classificato non prende nulla.
    """
    if "RegoleCrediti" not in wb.sheetnames:
        avvisi.append("foglio RegoleCrediti assente: uso i valori predefiniti")
        return {k: dict(v) for k, v in PREDEFINITE.items()}

    regole: dict = {}
    for r in leggi(wb["RegoleCrediti"]):
        categoria = str(r.get("Categoria") or "").strip()
        valore = intero(r.get("Valore"))
        if not categoria or valore is None:
            continue
        posizione = intero(r.get("Posizione")) or 1
        regole.setdefault(categoria, {})[posizione] = valore

    for categoria, valori in PREDEFINITE.items():
        if categoria not in regole:
            avvisi.append(f"regola '{categoria}' assente nel foglio: uso il predefinito")
            regole[categoria] = dict(valori)
    return regole


def valore(regole, categoria, posizione=1):
    """Valore di una regola, con ripiego sul predefinito se manca."""
    v = regole.get(categoria, {}).get(posizione)
    if v is not None:
        return v
    return PREDEFINITE.get(categoria, {}).get(posizione, 0)


def leggi(ws):
    """Righe del foglio come dizionari."""
    righe = [r for r in ws.iter_rows(values_only=True)
             if any(c is not None and str(c).strip() != "" for c in r)]
    if len(righe) < 2:
        return []
    intestazioni = [str(h).strip() if h is not None else "" for h in righe[0]]
    return [
        {k: v for k, v in zip(intestazioni, r) if k}
        for r in righe[1:]
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Compila CreditiAnnoNuovo")
    parser.add_argument("--anno", type=int, help="stagione da calcolare (default: tutte)")
    parser.add_argument("--prova", action="store_true",
                        help="mostra i conti senza salvare il file")
    args = parser.parse_args()

    if not XLSX.exists():
        print(f"ERRORE: non trovo {XLSX}")
        return 1

    wb = openpyxl.load_workbook(XLSX)

    avvisi: list[str] = []
    regole = carica_regole(wb, avvisi)
    base = valore(regole, "base")
    tetto_max = valore(regole, "tetto_max")
    tetto_min = valore(regole, "tetto_min")
    bonus_punteggio_pieno = valore(regole, "best_score")

    print(f"Regole: base {base}, tetti {tetto_min}-{tetto_max}, "
          f"miglior punteggio +{bonus_punteggio_pieno}, "
          f"coppa 1o +{valore(regole, 'coppa', 1)}")
    for a in avvisi:
        print(f"  ! {a}")

    squadre = {}
    for r in leggi(wb["AnagraficaSquadre"]):
        idx = intero(r.get("ID"))
        if idx is not None:
            squadre[idx] = str(r.get("Squadra") or "").strip()

    # Classifica: posizione e flag miglior punteggio, per anno
    classifica: dict[int, dict[int, dict]] = {}
    for r in leggi(wb["Classifiche"]):
        anno, idx = intero(r.get("Anno")), intero(r.get("ID"))
        if anno is None or idx is None:
            continue
        classifica.setdefault(anno, {})[idx] = {
            "posizione": intero(r.get("Posizione")),
            "miglior": bool(intero(r.get("Miglior Punteggio"), 0)),
        }

    # Coppa: posizione, per anno
    coppa: dict[int, dict[int, int]] = {}
    for r in leggi(wb["Coppa"]):
        anno, idx = intero(r.get("Anno")), intero(r.get("ID"))
        if anno is None or idx is None:
            continue
        coppa.setdefault(anno, {})[idx] = intero(r.get("Posizione"))

    # Residui di fine stagione
    residui: dict[int, dict[int, int]] = {}
    for r in leggi(wb["CreditiAsta"]):
        anno, idx = intero(r.get("Anno")), intero(r.get("ID"))
        if anno is None or idx is None:
            continue
        residui.setdefault(anno, {})[idx] = intero(r.get("CreditiFineAnno"), 0)

    anni = [args.anno] if args.anno else sorted(classifica.keys())
    if not anni:
        print("Nessuna stagione in Classifiche: niente da calcolare.")
        return 1

    ws = wb["CreditiAnnoNuovo"]
    esistenti = {}
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        anno, idx = intero(r[0]), intero(r[1])
        if anno is not None and idx is not None:
            esistenti[(anno, idx)] = i

    scritte = 0
    for anno in anni:
        print(f"\n=== Stagione {anno} → budget {anno + 1} ===")
        print(f"{'Squadra':<28}{'Pos':>4}{'Camp':>6}{'Coppa':>6}"
              f"{'Punt':>6}{'Resid':>7}{'Budget':>8}")
        print("-" * 65)

        for idx in sorted(squadre):
            riga_cl = classifica.get(anno, {}).get(idx)
            if riga_cl is None:
                continue

            posizione = riga_cl["posizione"]
            campionato = valore(regole, "classifica", posizione)
            punteggio = bonus_punteggio_pieno if riga_cl["miglior"] else 0

            pos_coppa = coppa.get(anno, {}).get(idx)
            bonus_coppa = valore(regole, "coppa", pos_coppa) if pos_coppa else 0

            resid = residui.get(anno, {}).get(idx, 0)

            grezzo = base + resid + campionato + bonus_coppa + punteggio
            budget = min(max(grezzo, tetto_min), tetto_max)

            segnale = ""
            if budget != grezzo:
                segnale = f"  <- tetto applicato (era {grezzo})"

            print(f"{squadre[idx][:27]:<28}{posizione:>4}{campionato:>+6}"
                  f"{bonus_coppa:>+6}{punteggio:>+6}{resid:>7}{budget:>8}{segnale}")

            valori = [anno, idx, squadre[idx], campionato, bonus_coppa,
                      punteggio, budget]
            r = esistenti.get((anno, idx))
            if r is None:
                ws.append(valori)
            else:
                for c, v in enumerate(valori, start=1):
                    ws.cell(row=r, column=c, value=v)
            scritte += 1

    if args.prova:
        print(f"\n[PROVA] {scritte} righe calcolate, file NON salvato.")
        return 0

    wb.save(XLSX)
    print(f"\nScritte {scritte} righe in {XLSX}")
    print("Ora lancia aggiorna.bat per portare i dati sul sito.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
