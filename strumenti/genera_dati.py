"""Converte Gestione.xlsx nel file di dati letto dal sito.

Uso:
    python strumenti/genera_dati.py
    python strumenti/genera_dati.py --xlsx "percorso\\del\\file.xlsx"

Legge i sei fogli di Gestione.xlsx e scrive `js/dati.js`, che le pagine del
sito caricano come un normale script.

Perché un .js e non un .json: un file JSON andrebbe letto con fetch(), che i
browser bloccano quando la pagina è aperta da file:// — il sito non sarebbe
più provabile in locale con un doppio click. Un .js che assegna una variabile
globale funziona ovunque, senza server.

Lo script NON calcola nulla: riporta i valori così come li hai scritti
nell'Excel. I premi e i malus non sono costanti (nel gestionale vivono nella
tabella `credit_rules` e cambiano per anno), quindi ricalcolarli qui
introdurrebbe valori che possono divergere da quelli veri.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Manca openpyxl. Installalo con:\n"
        "    pip install openpyxl\n"
        "oppure usa l'ambiente del gestionale:\n"
        '    d:\\Codice\\Fanta_Riconferme\\fantacalcio\\.venv\\Scripts\\python.exe '
        "strumenti/genera_dati.py"
    )

RADICE = Path(__file__).resolve().parent.parent
XLSX_PREDEFINITO = RADICE / "dati" / "Gestione.xlsx"
USCITA = RADICE / "js" / "dati.js"

DURATA_CONTRATTO = 4  # regolamento § 1
ETA_MAX_C = 21        # regolamento § 2
ETA_MAX_B = 25
MAX_CATEGORIA_A = 3
MAX_A_PIU_B = 6


def categoria_per_eta(anno_nascita, anno_riferimento):
    """Stessa regola di js/calcoli.js e del gestionale."""
    if anno_nascita is None or anno_riferimento is None:
        return "A"
    eta = anno_riferimento - anno_nascita
    if eta <= ETA_MAX_C:
        return "C"
    if eta <= ETA_MAX_B:
        return "B"
    return "A"


def calcola_penale(prezzo, anni_rimanenti):
    """Penale di svincolo — regolamento § 8. Come calcolaSvincolo() in JS."""
    if not anni_rimanenti or anni_rimanenti <= 0:
        return 0
    return math.ceil(prezzo * 0.10 * anni_rimanenti - 1e-9)


# --- Lettura dei fogli ---------------------------------------------------

def leggi_foglio(wb, nome: str) -> list[dict]:
    """Restituisce le righe di un foglio come dizionari, saltando le vuote."""
    if nome not in wb.sheetnames:
        print(f"  ! foglio '{nome}' assente, lo salto")
        return []

    ws = wb[nome]
    righe = [
        r for r in ws.iter_rows(values_only=True)
        if any(c is not None and str(c).strip() != "" for c in r)
    ]
    if len(righe) < 2:
        # Solo l'intestazione: è uno stato legittimo (Svincoli è vuoto finché
        # non ne fai uno), non un problema da segnalare.
        return []

    intestazioni = [str(h).strip() if h is not None else "" for h in righe[0]]
    risultato = []
    for riga in righe[1:]:
        voce = {}
        for chiave, valore in zip(intestazioni, riga):
            if chiave:
                voce[chiave] = valore
        risultato.append(voce)
    return risultato


def intero(valore, predefinito=None):
    """Converte in int i numeri che Excel salva come float (1.0 -> 1)."""
    if valore is None or (isinstance(valore, str) and not valore.strip()):
        return predefinito
    try:
        return int(float(valore))
    except (TypeError, ValueError):
        return predefinito


def numero(valore, predefinito=None):
    """Converte in numero mantenendo i decimali (i punteggi hanno il ,5)."""
    if valore is None or (isinstance(valore, str) and not valore.strip()):
        return predefinito
    try:
        f = float(valore)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return predefinito


def testo(valore, predefinito=""):
    if valore is None:
        return predefinito
    return str(valore).strip()


# --- Costruzione della struttura dati ------------------------------------

def costruisci(wb) -> dict:
    dati: dict = {
        "generatoIl": dt.date.today().isoformat(),
        "squadre": [],
        "nomiPerAnno": {},
        "contratti": [],
        "svincoli": [],
        "classifiche": {},
        "coppa": {},
        "creditiResidui": {},
        "creditiAnnoNuovo": {},
    }
    anni: set[int] = set()

    # Anagrafica: l'elenco delle squadre e di chi le allena
    print("  · AnagraficaSquadre")
    for riga in leggi_foglio(wb, "AnagraficaSquadre"):
        idx = intero(riga.get("ID"))
        if idx is None:
            continue
        dati["squadre"].append({
            "id": idx,
            "allenatore": testo(riga.get("Allenatore")),
            "squadra": testo(riga.get("Squadra")),
        })

    # Nome della squadra per ciascun anno (gestisce i cambi di nome)
    print("  · CambioNome")
    for riga in leggi_foglio(wb, "CambioNome"):
        idx, anno = intero(riga.get("ID")), intero(riga.get("Anno"))
        if idx is None or anno is None:
            continue
        anni.add(anno)
        dati["nomiPerAnno"].setdefault(str(anno), {})[str(idx)] = testo(riga.get("Nome"))

    # Contratti — l'anno di fine è sempre inizio + 3 e non si scrive
    print("  · Contratti")
    for riga in leggi_foglio(wb, "Contratti"):
        nome = testo(riga.get("Giocatore"))
        inizio = intero(riga.get("AnnoInizio"))
        # Le righe di esempio del file iniziale non finiscono nel sito
        if not nome or nome.upper().startswith("ESEMPIO") or inizio is None:
            continue

        fine = intero(riga.get("AnnoFine"), inizio + DURATA_CONTRATTO - 1)
        anni.add(inizio)
        dati["contratti"].append({
            "giocatore": nome,
            "annoNascita": intero(riga.get("AnnoNascita")),
            "idSquadra": intero(riga.get("IDSquadra")),
            "annoInizio": inizio,
            "annoFine": fine,
            "prezzo": intero(riga.get("PrezzoAcquisto"), 0),
            "attivo": bool(intero(riga.get("Attivo"), 1)),
        })

    # Svincoli — la penale si calcola dal contratto, salvo forzatura
    print("  · Svincoli")
    for riga in leggi_foglio(wb, "Svincoli"):
        nome = testo(riga.get("Giocatore"))
        anno_sv = intero(riga.get("AnnoSvincolo"))
        if not nome or nome.upper().startswith("ESEMPIO") or anno_sv is None:
            continue

        id_squadra = intero(riga.get("IDSquadra"))
        anni.add(anno_sv)

        # Cerca il contratto a cui si riferisce: stesso giocatore, stessa
        # squadra, e anno di svincolo dentro la durata.
        contratto = None
        for c in dati["contratti"]:
            if (c["giocatore"].strip().lower() == nome.strip().lower()
                    and c["idSquadra"] == id_squadra
                    and c["annoInizio"] <= anno_sv <= c["annoFine"]):
                contratto = c
                break

        penale = intero(riga.get("Penale"))
        forzata = penale is not None
        anni_rimanenti = None
        if contratto is not None:
            anni_rimanenti = contratto["annoFine"] - anno_sv + 1
            if penale is None:
                penale = calcola_penale(contratto["prezzo"], anni_rimanenti)

        dati["svincoli"].append({
            "giocatore": nome,
            "idSquadra": id_squadra,
            "anno": anno_sv,
            "penale": penale if penale is not None else 0,
            "penaleForzata": forzata,
            "anniRimanenti": anni_rimanenti,
            "prezzo": contratto["prezzo"] if contratto else None,
            "note": testo(riga.get("Note")),
            "contrattoTrovato": contratto is not None,
        })

    # Classifica di campionato
    print("  · Classifiche")
    for riga in leggi_foglio(wb, "Classifiche"):
        anno = intero(riga.get("Anno"))
        if anno is None:
            continue
        anni.add(anno)
        dati["classifiche"].setdefault(str(anno), []).append({
            "posizione": intero(riga.get("Posizione")),
            "id": intero(riga.get("ID")),
            "squadra": testo(riga.get("Squadra")),
            "punteggio": numero(riga.get("Punteggio")),
            "migliorPunteggio": bool(intero(riga.get("Miglior Punteggio"), 0)),
        })

    # Coppa
    print("  · Coppa")
    for riga in leggi_foglio(wb, "Coppa"):
        anno = intero(riga.get("Anno"))
        if anno is None:
            continue
        anni.add(anno)
        dati["coppa"].setdefault(str(anno), []).append({
            "posizione": intero(riga.get("Posizione")),
            "id": intero(riga.get("ID")),
            "squadra": testo(riga.get("Squadra")),
        })

    # Crediti avanzati a fine stagione
    print("  · CreditiAsta")
    for riga in leggi_foglio(wb, "CreditiAsta"):
        idx, anno = intero(riga.get("ID")), intero(riga.get("Anno"))
        if idx is None or anno is None:
            continue
        anni.add(anno)
        dati["creditiResidui"].setdefault(str(anno), {})[str(idx)] = {
            "crediti": intero(riga.get("CreditiFineAnno"), 0),
            "nome": testo(riga.get("Nome")),
        }

    # Budget della stagione successiva (compilato a mano nell'Excel)
    print("  · CreditiAnnoNuovo")
    for riga in leggi_foglio(wb, "CreditiAnnoNuovo"):
        anno = intero(riga.get("Anno"))
        if anno is None:
            continue
        anni.add(anno)
        dati["creditiAnnoNuovo"].setdefault(str(anno), []).append({
            "id": intero(riga.get("ID")),
            "squadra": testo(riga.get("Squadra")),
            "campionato": intero(riga.get("Crediti da Campionato")),
            "coppa": intero(riga.get("Crediti da Coppa")),
            "punteggio": intero(riga.get("Crediti da Punteggio")),
            "totale": intero(riga.get("Crediti per Anno Nuovo")),
        })

    # Ordinamenti stabili, così il file non cambia senza motivo fra due run
    dati["squadre"].sort(key=lambda s: s["id"])
    dati["contratti"].sort(
        key=lambda c: (c["idSquadra"] is None, c["idSquadra"] or 0, c["giocatore"])
    )
    dati["svincoli"].sort(key=lambda s: (-s["anno"], s["idSquadra"] or 0, s["giocatore"]))
    for elenco in dati["classifiche"].values():
        elenco.sort(key=lambda r: (r["posizione"] is None, r["posizione"]))
    for elenco in dati["coppa"].values():
        elenco.sort(key=lambda r: (r["posizione"] is None, r["posizione"]))
    for elenco in dati["creditiAnnoNuovo"].values():
        elenco.sort(key=lambda r: (r["id"] is None, r["id"]))

    dati["anni"] = sorted(anni, reverse=True)
    costruisci_albo_doro(dati)
    costruisci_stagioni(dati)
    return dati


def costruisci_albo_doro(dati: dict) -> None:
    """Vincitori stagione per stagione e medagliere complessivo.

    Si ricava da Classifiche e Coppa: non è un dato da compilare a mano, e
    ricalcolarlo a ogni generazione evita che resti indietro.
    """
    nomi = {s["id"]: s["squadra"] for s in dati["squadre"]}

    def nome(anno, idx, ripiego=""):
        per_anno = dati["nomiPerAnno"].get(str(anno), {})
        return per_anno.get(str(idx)) or ripiego or nomi.get(idx, f"Squadra {idx}")

    per_anno = []
    medaglie: dict = {}

    def conta(idx, chiave):
        v = medaglie.setdefault(idx, {"ori": 0, "argenti": 0, "bronzi": 0,
                                      "coppe": 0, "miglioriPunteggi": 0})
        v[chiave] += 1

    for anno_str, classifica in dati["classifiche"].items():
        anno = int(anno_str)
        podio = {r["posizione"]: r for r in classifica if r["posizione"] in (1, 2, 3)}
        migliore = next((r for r in classifica if r["migliorPunteggio"]), None)
        coppa = dati["coppa"].get(anno_str, [])
        vincitore_coppa = next((r for r in coppa if r["posizione"] == 1), None)

        voce = {"anno": anno}
        for posizione, chiave in ((1, "campione"), (2, "secondo"), (3, "terzo")):
            r = podio.get(posizione)
            if r:
                voce[chiave] = {"id": r["id"], "squadra": nome(anno, r["id"], r["squadra"]),
                                "punteggio": r["punteggio"]}
                conta(r["id"], {1: "ori", 2: "argenti", 3: "bronzi"}[posizione])
            else:
                voce[chiave] = None

        if vincitore_coppa:
            voce["coppa"] = {"id": vincitore_coppa["id"],
                             "squadra": nome(anno, vincitore_coppa["id"],
                                             vincitore_coppa["squadra"])}
            conta(vincitore_coppa["id"], "coppe")
        else:
            voce["coppa"] = None

        if migliore:
            voce["migliorPunteggio"] = {
                "id": migliore["id"],
                "squadra": nome(anno, migliore["id"], migliore["squadra"]),
                "punteggio": migliore["punteggio"],
            }
            conta(migliore["id"], "miglioriPunteggi")
        else:
            voce["migliorPunteggio"] = None

        per_anno.append(voce)

    per_anno.sort(key=lambda v: -v["anno"])

    medagliere = []
    for idx, v in medaglie.items():
        medagliere.append({"id": idx, "squadra": nomi.get(idx, f"Squadra {idx}"), **v})
    # Ordine olimpico: prima gli ori, poi argenti, poi bronzi, poi le coppe
    medagliere.sort(key=lambda m: (-m["ori"], -m["argenti"], -m["bronzi"],
                                   -m["coppe"], m["squadra"]))

    dati["alboDoro"] = {"perAnno": per_anno, "medagliere": medagliere}


def costruisci_stagioni(dati: dict) -> None:
    """Elenco delle stagioni selezionabili, e il budget di ciascuna.

    Il foglio CreditiAnnoNuovo è indicizzato sull'anno che si CHIUDE, ma quel
    budget si spende nell'asta dell'anno DOPO. Qui viene reindicizzato
    sull'anno in cui si usa, che è come lo cerca chi guarda la pagina: la
    stagione 2027 esiste proprio perché ha un budget, pur non essendo ancora
    stata giocata.
    """
    budget: dict = {}
    for anno_str, righe in dati["creditiAnnoNuovo"].items():
        anno_chiusura = int(anno_str)
        anno_uso = anno_chiusura + 1
        residui = dati["creditiResidui"].get(anno_str, {})
        voci = []
        for r in righe:
            res = residui.get(str(r["id"]))
            voci.append({
                "id": r["id"],
                "squadra": r["squadra"],
                "campionato": r["campionato"],
                "coppa": r["coppa"],
                "punteggio": r["punteggio"],
                "residui": res["crediti"] if res else None,
                "totale": r["totale"],
                "daStagione": anno_chiusura,
            })
        if voci:
            budget[str(anno_uso)] = voci

    dati["budget"] = budget

    anni = set(int(a) for a in dati["classifiche"])
    anni |= set(int(a) for a in budget)
    anni |= set(int(a) for a in dati["coppa"])

    dati["stagioni"] = [
        {"anno": a,
         "giocata": str(a) in dati["classifiche"],
         "haBudget": str(a) in budget}
        for a in sorted(anni, reverse=True)
    ]


def controlli(dati: dict) -> list[str]:
    """Segnala incoerenze senza correggerle: le correzioni si fanno nell'Excel."""
    avvisi = []
    id_noti = {s["id"] for s in dati["squadre"]}

    for anno, classifica in dati["classifiche"].items():
        posizioni = [r["posizione"] for r in classifica if r["posizione"] is not None]
        if len(posizioni) != len(set(posizioni)):
            avvisi.append(f"{anno}: posizioni duplicate in classifica")

        for riga in classifica:
            if riga["id"] not in id_noti:
                avvisi.append(f"{anno}: ID {riga['id']} in classifica ma non in anagrafica")

        con_punteggio = [r for r in classifica if r["punteggio"] is not None]
        migliori = [r for r in classifica if r["migliorPunteggio"]]
        if len(migliori) > 1:
            avvisi.append(f"{anno}: più di una squadra segnata come miglior punteggio")
        if migliori and con_punteggio:
            massimo = max(r["punteggio"] for r in con_punteggio)
            if migliori[0]["punteggio"] != massimo:
                avvisi.append(
                    f"{anno}: '{migliori[0]['squadra']}' è segnata come miglior punteggio "
                    f"({migliori[0]['punteggio']}) ma il punteggio più alto è {massimo}"
                )

    for anno, righe in dati["creditiAnnoNuovo"].items():
        vuote = [r["squadra"] for r in righe if r["totale"] is None]
        if vuote:
            avvisi.append(
                f"{anno}: 'Crediti per Anno Nuovo' non compilato per {len(vuote)} squadre "
                "(il sito mostrerà un trattino)"
            )

    # --- Contratti ---
    attivi_per_giocatore: dict[str, int] = {}

    for c in dati["contratti"]:
        nome = c["giocatore"]
        if c["idSquadra"] not in id_noti:
            avvisi.append(
                f"Contratti: '{nome}' ha IDSquadra {c['idSquadra']}, che non "
                "esiste in AnagraficaSquadre"
            )
        if c["annoNascita"] is None:
            avvisi.append(
                f"Contratti: '{nome}' senza AnnoNascita — verrà trattato come "
                "categoria A, e il conteggio degli slot potrebbe risultare falsato"
            )
        if not c["prezzo"] or c["prezzo"] <= 0:
            avvisi.append(
                f"Contratti: '{nome}' ha PrezzoAcquisto {c['prezzo']} — "
                "i costi verranno calcolati male"
            )
        if c["attivo"]:
            chiave = nome.strip().lower()
            attivi_per_giocatore[chiave] = attivi_per_giocatore.get(chiave, 0) + 1

    for nome, quanti in attivi_per_giocatore.items():
        if quanti > 1:
            avvisi.append(
                f"Contratti: risultano {quanti} contratti attivi per '{nome}' — "
                "o è un doppione, o sono due omonimi (in tal caso distinguili "
                "nel nome)"
            )

    # --- Svincoli ---
    attivi = {
        (c["giocatore"].strip().lower(), c["idSquadra"])
        for c in dati["contratti"] if c["attivo"]
    }
    for s in dati["svincoli"]:
        if not s["contrattoTrovato"]:
            avvisi.append(
                f"Svincoli: per '{s['giocatore']}' nel {s['anno']} non trovo un "
                "contratto corrispondente — controlla nome, IDSquadra e che "
                "l'anno cada dentro la durata del contratto"
            )
        if (s["giocatore"].strip().lower(), s["idSquadra"]) in attivi:
            avvisi.append(
                f"Svincoli: '{s['giocatore']}' risulta svincolato nel {s['anno']} "
                "ma il suo contratto è ancora Attivo = 1 — mettilo a 0"
            )
        if s["penaleForzata"]:
            avvisi.append(
                f"Svincoli: la penale di '{s['giocatore']}' ({s['penale']}) è "
                "scritta a mano, non calcolata. Lascia la cella vuota se vuoi "
                "il valore da regolamento"
            )

    # --- Slot per squadra e anno ---
    per_squadra_anno: dict = {}

    for c in dati["contratti"]:
        if not c["attivo"]:
            continue
        for anno in range(c["annoInizio"], c["annoFine"] + 1):
            cat = categoria_per_eta(c["annoNascita"], anno)
            chiave = (c["idSquadra"], anno)
            conta = per_squadra_anno.setdefault(chiave, {"A": 0, "B": 0, "C": 0})
            conta[cat] += 1

    nomi_squadre = {s["id"]: s["squadra"] for s in dati["squadre"]}
    # L'id può essere None quando il nome della squadra non è stato riconosciuto:
    # quel caso è già segnalato sopra, qui basta non far esplodere l'ordinamento.
    for (id_sq, anno), conta in sorted(
        per_squadra_anno.items(),
        key=lambda voce: (voce[0][0] is None, voce[0][0] or 0, voce[0][1]),
    ):
        squadra = nomi_squadre.get(id_sq, f"squadra {id_sq}")
        if conta["A"] > MAX_CATEGORIA_A:
            avvisi.append(
                f"{anno}: {squadra} ha {conta['A']} giocatori in categoria A "
                f"(massimo {MAX_CATEGORIA_A})"
            )
        if conta["A"] + conta["B"] > MAX_A_PIU_B:
            avvisi.append(
                f"{anno}: {squadra} ha {conta['A'] + conta['B']} giocatori in A+B "
                f"(massimo {MAX_A_PIU_B})"
            )

    return avvisi


def scrivi(dati: dict, destinazione: Path) -> None:
    corpo = json.dumps(dati, ensure_ascii=False, indent=2, sort_keys=False)
    destinazione.parent.mkdir(parents=True, exist_ok=True)
    destinazione.write_text(
        "/* GENERATO AUTOMATICAMENTE — non modificare a mano.\n"
        " * Prodotto da strumenti/genera_dati.py a partire da Gestione.xlsx.\n"
        " * Per aggiornarlo: modifica l'Excel e rilancia lo script.\n"
        " */\n"
        "window.DATI = " + corpo + ";\n",
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Genera js/dati.js da Gestione.xlsx")
    parser.add_argument("--xlsx", type=Path, default=XLSX_PREDEFINITO,
                        help="percorso del file Excel di partenza")
    parser.add_argument("--out", type=Path, default=USCITA,
                        help="file da scrivere")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERRORE: non trovo {args.xlsx}")
        return 1

    print(f"Leggo {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    dati = costruisci(wb)

    scrivi(dati, args.out)

    print(f"\nScritto {args.out}")
    print(f"  {len(dati['squadre'])} squadre")
    n_attivi = sum(1 for c in dati["contratti"] if c["attivo"])
    print(f"  {len(dati['contratti'])} contratti ({n_attivi} attivi)")
    if dati["svincoli"]:
        penali = sum(s["penale"] for s in dati["svincoli"])
        print(f"  {len(dati['svincoli'])} svincoli, {penali} crediti di penali")
    stagioni = dati.get("stagioni", [])
    if stagioni:
        etichette = [
            f"{s['anno']}" + ("" if s["giocata"] else " (da giocare)")
            for s in stagioni
        ]
        print(f"  {len(stagioni)} stagioni: {', '.join(etichette)}")
    albo = dati.get("alboDoro", {}).get("perAnno", [])
    if albo:
        c = albo[0].get("campione")
        print(f"  albo d'oro: {len(albo)} edizioni"
              + (f", ultimo campione {c['squadra']}" if c else ""))

    avvisi = controlli(dati)
    if avvisi:
        print("\nDa controllare nell'Excel:")
        for a in avvisi:
            print(f"  ! {a}")
    else:
        print("\nNessuna incoerenza rilevata.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
